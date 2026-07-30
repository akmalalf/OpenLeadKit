"""Safe Excel export with optional custom workbook support."""

from __future__ import annotations

import hashlib
import re
import shutil
import uuid
import zipfile
from collections.abc import Iterable
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from openleadkit.exceptions import ExportError, WorkbookCompatibilityError
from openleadkit.services.normalization import (
    extract_domain,
    normalize_business_name,
    normalize_phone,
)

LOGICAL_COLUMNS = (
    "Imported At",
    "Source",
    "Search Query",
    "Business Name",
    "Category",
    "City",
    "Address",
    "Source URL",
    "Website URL",
    "Phone",
    "Email",
    "Instagram",
    "Google Rating",
    "Review Count",
    "Opening Hours",
    "Latitude",
    "Longitude",
    "Scraper/Method",
    "Batch ID",
    "Raw Notes",
)
REQUIRED_COLUMNS = {"Business Name", "Source URL", "Batch ID"}
EXTENSION_PATTERN = re.compile(rb"<extLst(?:\s[^>]*)?>.*?</extLst>", re.DOTALL)
XR_NAMESPACE = b' xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"'
GENERATED_HEADER_FILL = PatternFill("solid", fgColor="246B49")
GENERATED_HEADER_FONT = Font(color="FFFFFF", bold=True)
GENERATED_COLUMN_WIDTHS = {
    "Imported At": 20,
    "Source": 18,
    "Search Query": 28,
    "Business Name": 32,
    "Category": 22,
    "City": 20,
    "Address": 42,
    "Source URL": 42,
    "Website URL": 36,
    "Phone": 20,
    "Email": 30,
    "Instagram": 24,
    "Google Rating": 15,
    "Review Count": 15,
    "Opening Hours": 28,
    "Latitude": 14,
    "Longitude": 14,
    "Scraper/Method": 34,
    "Batch ID": 32,
    "Raw Notes": 48,
}


def _key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def raw_import_sheet(workbook: Any) -> Worksheet:
    for name in workbook.sheetnames:
        if name.casefold().strip() == "raw import":
            return workbook[name]
    raise WorkbookCompatibilityError("The 'Raw Import' worksheet was not found")


def detect_header_row(sheet: Worksheet, scan_limit: int = 50) -> int:
    expected = {_key(column) for column in LOGICAL_COLUMNS}
    best_row, best_score = 0, 0
    for row in range(1, min(sheet.max_row, scan_limit) + 1):
        values = {_key(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1)}
        score = len(expected & values)
        if score > best_score:
            best_row, best_score = row, score
    if best_score < 3:
        raise WorkbookCompatibilityError("The Raw Import header row could not be identified")
    return best_row


def header_mapping(sheet: Worksheet, header_row: int) -> dict[str, int]:
    actual = {
        _key(sheet.cell(header_row, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(header_row, column).value is not None
    }
    return {
        logical: actual[_key(logical)] for logical in LOGICAL_COLUMNS if _key(logical) in actual
    }


def first_usable_data_row(sheet: Worksheet, header_row: int, mapping: dict[str, int]) -> int:
    name_column = mapping.get("Business Name")
    if name_column is None:
        raise WorkbookCompatibilityError("The Business Name column was not found")
    row = header_row + 1
    while row <= max(sheet.max_row, header_row + 1):
        if sheet.cell(row, name_column).value in (None, ""):
            return row
        row += 1
    return row


@dataclass(frozen=True)
class WorkbookInspection:
    worksheet_names: tuple[str, ...]
    raw_import_name: str
    header_row: int
    columns: tuple[str, ...]
    first_data_row: int
    first_empty_row: int
    formula_cells: tuple[str, ...]
    styled_rows: tuple[int, ...]


def inspect_workbook(path: Path) -> WorkbookInspection:
    if not path.is_file():
        raise WorkbookCompatibilityError(f"Workbook not found: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = raw_import_sheet(workbook)
        row = detect_header_row(sheet)
        mapping = header_mapping(sheet, row)
        missing = REQUIRED_COLUMNS - mapping.keys()
        if missing:
            raise WorkbookCompatibilityError(
                "Required columns were not found: " + ", ".join(sorted(missing))
            )
        formulas = tuple(
            cell.coordinate for line in sheet.iter_rows() for cell in line if cell.data_type == "f"
        )
        styled_rows = tuple(
            row_number
            for row_number in range(row + 1, sheet.max_row + 1)
            if any(sheet.cell(row_number, col).has_style for col in mapping.values())
        )
        return WorkbookInspection(
            worksheet_names=tuple(workbook.sheetnames),
            raw_import_name=sheet.title,
            header_row=row,
            columns=tuple(
                str(sheet.cell(row, col).value)
                for col in range(1, sheet.max_column + 1)
                if sheet.cell(row, col).value is not None
            ),
            first_data_row=row + 1,
            first_empty_row=first_usable_data_row(sheet, row, mapping),
            formula_cells=formulas,
            styled_rows=styled_rows,
        )
    finally:
        workbook.close()


def read_exported_rows(
    path: Path,
    batch_id: str,
    *,
    offset: int = 0,
    limit: int = 25,
) -> tuple[dict[str, Any], ...]:
    """Read one bounded page of exact rows from a previously exported workbook."""
    if offset < 0:
        raise ValueError("The export row offset cannot be negative")
    if not 1 <= limit <= 100:
        raise ValueError("The export row limit must be between 1 and 100")
    if not path.is_file():
        raise WorkbookCompatibilityError(f"Workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = raw_import_sheet(workbook)
        header_row = detect_header_row(sheet)
        mapping = header_mapping(sheet, header_row)
        missing = REQUIRED_COLUMNS - mapping.keys()
        if missing:
            raise WorkbookCompatibilityError(
                "Required columns were not found: " + ", ".join(sorted(missing))
            )

        batch_index = mapping["Batch ID"] - 1
        max_column = max(mapping.values())
        matched = 0
        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(
            min_row=header_row + 1,
            min_col=1,
            max_col=max_column,
            values_only=True,
        ):
            if values[batch_index] != batch_id:
                continue
            if matched < offset:
                matched += 1
                continue
            rows.append(
                {
                    logical_name: values[column_number - 1]
                    for logical_name, column_number in mapping.items()
                }
            )
            if len(rows) >= limit:
                break
        return tuple(rows)
    finally:
        workbook.close()


def create_default_workbook(path: Path) -> None:
    """Create a neutral OpenLeadKit workbook without relying on a private template."""
    workbook = Workbook()
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        workbook.close()
        raise ExportError("OpenLeadKit could not create the Raw Import worksheet")
    sheet.title = "Raw Import"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 24
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(LOGICAL_COLUMNS))}1"

    for column_number, logical_name in enumerate(LOGICAL_COLUMNS, 1):
        header = sheet.cell(1, column_number, logical_name)
        header.fill = GENERATED_HEADER_FILL
        header.font = GENERATED_HEADER_FONT
        header.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(column_number)].width = GENERATED_COLUMN_WIDTHS[
            logical_name
        ]

    sheet.cell(2, 1).number_format = "yyyy-mm-dd hh:mm:ss"
    sheet.cell(2, 16).number_format = "0.000000"
    sheet.cell(2, 17).number_format = "0.000000"
    workbook.properties.creator = "OpenLeadKit"
    workbook.properties.title = "OpenLeadKit lead export"
    workbook.save(path)
    workbook.close()


@dataclass(frozen=True)
class ExportRecord:
    business_id: uuid.UUID
    business_name: str
    category: str
    city: str | None
    address: str | None
    source_url: str
    website_url: str | None
    phone: str | None
    email: str | None
    instagram: str | None
    opening_hours: str | None
    latitude: float
    longitude: float
    search_query: str
    raw_notes: str | None


def generate_batch_id(records: Iterable[ExportRecord], now: datetime) -> str:
    ids = "|".join(sorted(str(record.business_id) for record in records))
    digest = hashlib.sha256(ids.encode()).hexdigest()[:8].upper()
    return f"OLK-{now:%Y%m%d-%H%M%S}-{digest}"


def row_mapping(record: ExportRecord, imported_at: datetime, batch_id: str) -> dict[str, Any]:
    excel_datetime = (
        imported_at.astimezone().replace(tzinfo=None)
        if imported_at.tzinfo is not None
        else imported_at
    )
    notes = " | ".join(
        part
        for part in (
            record.raw_notes,
            "Atribusi: © OpenStreetMap contributors (ODbL)",
        )
        if part
    )
    return {
        "Imported At": excel_datetime,
        "Source": "OpenStreetMap",
        "Search Query": record.search_query,
        "Business Name": record.business_name,
        "Category": record.category,
        "City": record.city,
        "Address": record.address,
        "Source URL": record.source_url,
        "Website URL": record.website_url,
        "Phone": record.phone,
        "Email": record.email,
        "Instagram": record.instagram,
        "Google Rating": None,
        "Review Count": None,
        "Opening Hours": record.opening_hours,
        "Latitude": record.latitude,
        "Longitude": record.longitude,
        "Scraper/Method": "OpenLeadKit – OpenStreetMap Overpass",  # noqa: RUF001
        "Batch ID": batch_id,
        "Raw Notes": notes,
    }


@dataclass(frozen=True)
class ExistingKeys:
    source_urls: frozenset[str]
    domains: frozenset[str]
    phones: frozenset[str]
    names_cities: frozenset[tuple[str, str]]


def existing_workbook_keys(
    sheet: Worksheet, header_row: int, mapping: dict[str, int]
) -> ExistingKeys:
    source_urls: set[str] = set()
    domains: set[str] = set()
    phones: set[str] = set()
    names_cities: set[tuple[str, str]] = set()
    for row in range(header_row + 1, sheet.max_row + 1):

        def value(name: str, row_number: int = row) -> Any:
            return sheet.cell(row_number, mapping[name]).value if name in mapping else None

        source = str(value("Source URL") or "").strip()
        if source:
            source_urls.add(source)
        if domain := extract_domain(str(value("Website URL") or "")):
            domains.add(domain)
        if phone := normalize_phone(str(value("Phone") or "")):
            phones.add(phone)
        name = normalize_business_name(str(value("Business Name") or ""))
        city = str(value("City") or "").casefold().strip()
        if name and city:
            names_cities.add((name, city))
    return ExistingKeys(
        frozenset(source_urls), frozenset(domains), frozenset(phones), frozenset(names_cities)
    )


def record_exists(record: ExportRecord, keys: ExistingKeys) -> bool:
    if record.source_url and record.source_url in keys.source_urls:
        return True
    if record.website_url and extract_domain(record.website_url) in keys.domains:
        return True
    if record.phone and normalize_phone(record.phone) in keys.phones:
        return True
    return (
        normalize_business_name(record.business_name),
        (record.city or "").casefold().strip(),
    ) in keys.names_cities


def keys_with_record(keys: ExistingKeys, record: ExportRecord) -> ExistingKeys:
    """Return workbook keys extended with one newly accepted export record."""
    source_urls = set(keys.source_urls)
    domains = set(keys.domains)
    phones = set(keys.phones)
    names_cities = set(keys.names_cities)
    if record.source_url:
        source_urls.add(record.source_url)
    if domain := extract_domain(record.website_url):
        domains.add(domain)
    if phone := normalize_phone(record.phone):
        phones.add(phone)
    name = normalize_business_name(record.business_name)
    city = (record.city or "").casefold().strip()
    if name and city:
        names_cities.add((name, city))
    return ExistingKeys(
        source_urls=frozenset(source_urls),
        domains=frozenset(domains),
        phones=frozenset(phones),
        names_cities=frozenset(names_cities),
    )


def _restore_unsupported_sheet_extensions(source: Path, output: Path) -> None:
    """Restore worksheet extension lists that openpyxl cannot round-trip.

    Current Excel templates can store modern data validation in ``extLst``. Openpyxl warns
    that it removes those extensions. The workbook order is preserved, so restoring the exact
    source fragments after the supported edit retains those validations without interpreting
    or mutating them.
    """
    fragments: dict[str, bytes] = {}
    with zipfile.ZipFile(source) as source_archive:
        for name in source_archive.namelist():
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                continue
            match = EXTENSION_PATTERN.search(source_archive.read(name))
            if match:
                fragment = match.group(0).replace(b"<extLst>", b"<extLst" + XR_NAMESPACE + b">", 1)
                fragments[name] = fragment
    if not fragments:
        return
    temporary = output.with_suffix(".extensions.tmp.xlsx")
    with (
        zipfile.ZipFile(output) as output_archive,
        zipfile.ZipFile(temporary, "w") as rewritten,
    ):
        for item in output_archive.infolist():
            data = output_archive.read(item.filename)
            restored_fragment = fragments.get(item.filename)
            if restored_fragment:
                if EXTENSION_PATTERN.search(data):
                    data = EXTENSION_PATTERN.sub(restored_fragment, data, count=1)
                else:
                    data = data.replace(b"</worksheet>", restored_fragment + b"</worksheet>")
            rewritten.writestr(item, data)
    temporary.replace(output)


@dataclass(frozen=True)
class ExportResult:
    output_path: Path
    batch_id: str
    exported_ids: tuple[uuid.UUID, ...]
    skipped_existing: int
    skipped_invalid: int


def export_workbook(
    source: Path | None,
    output_dir: Path,
    records: list[ExportRecord],
    *,
    now: datetime,
) -> ExportResult:
    inspection = inspect_workbook(source) if source is not None else None
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = (
        f"Website_Lead_Funnel_CRM_{now:%Y%m%d_%H%M%S}.xlsx"
        if source is not None
        else f"OpenLeadKit_Leads_{now:%Y%m%d_%H%M%S}.xlsx"
    )
    output = output_dir / output_name
    if output.exists():
        raise ExportError("The output filename already exists; retry in one second")
    batch_id = generate_batch_id(records, now)
    try:
        if source is None:
            create_default_workbook(output)
            inspection = inspect_workbook(output)
        else:
            shutil.copy2(source, output)
        if inspection is None:  # Defensive guard for type checkers and unexpected failures.
            raise ExportError("Workbook inspection was not available")
        workbook = load_workbook(output)
        sheet = raw_import_sheet(workbook)
        mapping = header_mapping(sheet, inspection.header_row)
        keys = existing_workbook_keys(sheet, inspection.header_row, mapping)
        row = first_usable_data_row(sheet, inspection.header_row, mapping)
        exported: list[uuid.UUID] = []
        skipped_existing = 0
        skipped_invalid = 0
        for record in records:
            if not record.business_name or not record.source_url:
                skipped_invalid += 1
                continue
            if record_exists(record, keys):
                skipped_existing += 1
                continue
            values = row_mapping(record, now, batch_id)
            template_row = max(inspection.header_row + 1, row - 1)
            for logical, column in mapping.items():
                target = sheet.cell(row, column)
                template = sheet.cell(template_row, column)
                if template.has_style and isinstance(target, Cell) and isinstance(template, Cell):
                    target.font = copy(template.font)  # type: ignore[assignment]
                    target.fill = copy(template.fill)  # type: ignore[assignment]
                    target.border = copy(template.border)  # type: ignore[assignment]
                    target.alignment = copy(template.alignment)  # type: ignore[assignment]
                    target.protection = copy(template.protection)  # type: ignore[assignment]
                if template.number_format:
                    target.number_format = template.number_format
                target.value = values.get(logical)
            exported.append(record.business_id)
            keys = keys_with_record(keys, record)
            row += 1
        if source is None:
            final_row = max(inspection.header_row, row - 1)
            final_column = get_column_letter(max(mapping.values()))
            sheet.auto_filter.ref = f"A{inspection.header_row}:{final_column}{final_row}"
        workbook.save(output)
        workbook.close()
        if source is not None:
            _restore_unsupported_sheet_extensions(source, output)

        verification = load_workbook(output, read_only=True, data_only=False)
        verify_sheet = raw_import_sheet(verification)
        verify_mapping = header_mapping(verify_sheet, inspection.header_row)
        batch_column = verify_mapping["Batch ID"]
        found = sum(
            1
            for row_number in range(inspection.header_row + 1, verify_sheet.max_row + 1)
            if verify_sheet.cell(row_number, batch_column).value == batch_id
        )
        verification.close()
        if found != len(exported):
            raise ExportError("Workbook verification failed after saving")
    except Exception as exc:
        if output.exists():
            output.unlink()
        if isinstance(exc, (ExportError, WorkbookCompatibilityError)):
            raise
        raise ExportError("Workbook export failed") from exc
    return ExportResult(
        output_path=output,
        batch_id=batch_id,
        exported_ids=tuple(exported),
        skipped_existing=skipped_existing,
        skipped_invalid=skipped_invalid,
    )
