import hashlib
import uuid
import zipfile
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from openleadkit.services.excel_exporter import (
    LOGICAL_COLUMNS,
    ExportRecord,
    _restore_unsupported_sheet_extensions,
    export_workbook,
    generate_batch_id,
    inspect_workbook,
    read_exported_rows,
    record_exists,
    row_mapping,
)


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard["A1"] = "=1+1"
    raw = workbook.create_sheet("Raw Import")
    raw.freeze_panes = "A6"
    raw.row_dimensions[5].height = 24
    for column, header in enumerate(LOGICAL_COLUMNS, 1):
        raw.cell(5, column, header)
        raw.cell(6, column).fill = PatternFill("solid", fgColor="EAF2EC")
    raw.auto_filter.ref = "A5:T500"
    workbook.create_sheet("Guide")
    workbook.save(path)


def export_record(identifier: int = 1, **changes: object) -> ExportRecord:
    values = {
        "business_id": uuid.UUID(int=identifier),
        "business_name": "Arunika Coffee",
        "category": "Cafe",
        "city": "London",
        "address": "21 Market Street, London",
        "source_url": f"https://www.openstreetmap.org/node/{identifier}",
        "website_url": "https://arunika.example",
        "phone": "+442071234567",
        "email": "hello@arunika.example",
        "instagram": "arunikacoffee",
        "opening_hours": "Mo-Su 08:00-22:00",
        "latitude": -6.921,
        "longitude": 107.607,
        "search_query": "Cafe — London",
        "raw_notes": "Qualification: High",
    }
    values.update(changes)
    return ExportRecord(**values)


def test_inspection_detects_actual_layout(tmp_path: Path) -> None:
    path = tmp_path / "crm.xlsx"
    build_workbook(path)
    result = inspect_workbook(path)
    assert result.worksheet_names == ("Dashboard", "Raw Import", "Guide")
    assert result.header_row == 5
    assert result.columns == LOGICAL_COLUMNS
    assert result.first_data_row == 6
    assert result.first_empty_row == 6
    assert "A1" not in result.formula_cells  # formulas on other sheets are outside Raw Import


def test_batch_id_and_row_mapping() -> None:
    now = datetime(2026, 7, 29, 12, 34, 56, tzinfo=UTC)
    record = export_record()
    expected_digest = hashlib.sha256(str(record.business_id).encode()).hexdigest()[:8].upper()
    assert generate_batch_id([record], now) == f"OLK-20260729-123456-{expected_digest}"
    row = row_mapping(record, now, "BATCH")
    assert row["Google Rating"] is None
    assert row["Review Count"] is None
    assert row["Scraper/Method"] == "OpenLeadKit – OpenStreetMap Overpass"  # noqa: RUF001
    assert "OpenStreetMap contributors" in row["Raw Notes"]


def test_export_modifies_only_raw_import_and_preserves_source(tmp_path: Path) -> None:
    source = tmp_path / "crm.xlsx"
    output_dir = tmp_path / "exports"
    build_workbook(source)
    original_bytes = source.read_bytes()
    now = datetime(2026, 7, 29, 12, 0, 0, tzinfo=UTC)
    result = export_workbook(source, output_dir, [export_record()], now=now)
    assert source.read_bytes() == original_bytes
    assert result.output_path.exists()
    workbook = load_workbook(result.output_path, data_only=False)
    assert workbook.sheetnames == ["Dashboard", "Raw Import", "Guide"]
    assert workbook["Dashboard"]["A1"].value == "=1+1"
    raw = workbook["Raw Import"]
    assert raw["D6"].value == "Arunika Coffee"
    assert raw["S6"].value == result.batch_id
    assert raw.freeze_panes == "A6"
    assert raw["D6"].fill.fgColor.rgb.endswith("EAF2EC")
    workbook.close()


def test_export_creates_standalone_workbook_without_source(tmp_path: Path) -> None:
    output_dir = tmp_path / "exports"
    now = datetime(2026, 7, 29, 12, 0, 0, tzinfo=UTC)
    result = export_workbook(None, output_dir, [export_record()], now=now)

    assert result.output_path.name == "OpenLeadKit_Leads_20260729_120000.xlsx"
    workbook = load_workbook(result.output_path, data_only=False)
    assert workbook.sheetnames == ["Raw Import"]
    raw = workbook["Raw Import"]
    assert tuple(raw.cell(1, column).value for column in range(1, 21)) == LOGICAL_COLUMNS
    assert raw["D2"].value == "Arunika Coffee"
    assert raw["S2"].value == result.batch_id
    assert raw.freeze_panes == "A2"
    assert raw.auto_filter.ref == "A1:T2"
    assert raw["A1"].font.bold
    assert raw["A1"].fill.fgColor.rgb.endswith("246B49")
    assert raw["A2"].number_format == "yyyy-mm-dd hh:mm:ss"
    assert raw["P2"].number_format == "0.000000"
    workbook.close()


def test_read_exported_rows_returns_one_exact_bounded_page(tmp_path: Path) -> None:
    records = [
        export_record(
            identifier,
            business_name=f"Business {identifier}",
            city=f"City {identifier}",
            website_url=f"https://business-{identifier}.example",
            phone=f"+44207123000{identifier}",
        )
        for identifier in range(1, 4)
    ]
    result = export_workbook(
        None,
        tmp_path / "exports",
        records,
        now=datetime(2026, 7, 29, 12, 0, 0, tzinfo=UTC),
    )

    rows = read_exported_rows(result.output_path, result.batch_id, offset=1, limit=1)

    assert len(rows) == 1
    assert rows[0]["Business Name"] == "Business 2"
    assert rows[0]["City"] == "City 2"
    assert rows[0]["Batch ID"] == result.batch_id


def test_read_exported_rows_rejects_unbounded_requests(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="offset"):
        read_exported_rows(tmp_path / "missing.xlsx", "BATCH", offset=-1)
    with pytest.raises(ValueError, match="limit"):
        read_exported_rows(tmp_path / "missing.xlsx", "BATCH", limit=101)


def test_export_skips_existing_records(tmp_path: Path) -> None:
    source = tmp_path / "crm.xlsx"
    output_dir = tmp_path / "exports"
    build_workbook(source)
    now = datetime(2026, 7, 29, 12, 0, 0, tzinfo=UTC)
    first = export_workbook(source, output_dir, [export_record()], now=now)
    second = export_workbook(
        first.output_path,
        output_dir,
        [
            export_record(),
            export_record(
                2,
                business_name="Northern Lens Studio",
                city="Manchester",
                website_url="https://northern-lens.example",
                phone="+441615550000",
            ),
        ],
        now=datetime(2026, 7, 29, 12, 0, 1, tzinfo=UTC),
    )
    assert second.skipped_existing == 1
    assert len(second.exported_ids) == 1


def test_export_skips_duplicates_within_the_same_batch(tmp_path: Path) -> None:
    source = tmp_path / "crm.xlsx"
    build_workbook(source)
    result = export_workbook(
        source,
        tmp_path / "exports",
        [export_record(), export_record(2)],
        now=datetime(2026, 7, 29, 12, 0, 0, tzinfo=UTC),
    )
    assert result.exported_ids == (uuid.UUID(int=1),)
    assert result.skipped_existing == 1
    workbook = load_workbook(result.output_path, read_only=True)
    assert workbook["Raw Import"].max_row == 6
    workbook.close()


def test_existing_record_checks_are_conservative(tmp_path: Path) -> None:
    source = tmp_path / "crm.xlsx"
    build_workbook(source)
    first = export_workbook(
        source,
        tmp_path / "exports",
        [export_record()],
        now=datetime(2026, 7, 29, 1, 0, 0, tzinfo=UTC),
    )
    workbook = load_workbook(first.output_path)
    from openleadkit.services.excel_exporter import (
        detect_header_row,
        existing_workbook_keys,
        header_mapping,
        raw_import_sheet,
    )

    sheet = raw_import_sheet(workbook)
    header = detect_header_row(sheet)
    keys = existing_workbook_keys(sheet, header, header_mapping(sheet, header))
    assert record_exists(export_record(99), keys)  # same domain/phone/name+city
    assert not record_exists(
        export_record(
            99,
            business_name="Northern Lens Studio",
            city="Manchester",
            website_url="https://northern-lens.example",
            phone="+441615550000",
        ),
        keys,
    )
    workbook.close()


def test_unsupported_excel_extensions_are_restored(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    extension = (
        b'<extLst><ext uri="validation" '
        b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
        b'<x14:dataValidations count="1"/></ext></extLst>'
    )
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"<worksheet>" + extension + b"</worksheet>")
    with zipfile.ZipFile(output, "w") as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"<worksheet></worksheet>")
        archive.writestr("xl/workbook.xml", b"<workbook/>")
    _restore_unsupported_sheet_extensions(source, output)
    with zipfile.ZipFile(output) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")
        assert b"x14:dataValidations" in worksheet
        assert b"xmlns:xr=" in worksheet
        assert archive.read("xl/workbook.xml") == b"<workbook/>"
